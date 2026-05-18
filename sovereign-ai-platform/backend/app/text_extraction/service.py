from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


def extract_text(path: Path, content_type: str | None = None) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        reader = PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    if suffix in {".docx", ".doc"}:
        doc = DocxDocument(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if suffix in {".xlsx", ".xls"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        rows: list[str] = []
        for sheet in workbook.worksheets:
            rows.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell) for cell in row if cell is not None]
                if values:
                    rows.append(" | ".join(values))
        return "\n".join(rows)
    return path.read_text(encoding="utf-8", errors="ignore")

